# -*- coding: utf-8 -*-
"""店舗xlsx → サイト(stores/index.html) 同期ツール

使い方:
  python -X utf8 scripts/sync_stores.py            # 差分レポートのみ（何も変更しない）
  python -X utf8 scripts/sync_stores.py --apply    # 時給変更をindex.htmlに反映
  python -X utf8 scripts/sync_stores.py --apply --commit  # 反映＋git commit/push

やること:
  1. xlsx（context/一次情報/店舗情報まとめ最新.xlsx）を読む
  2. stores/index.html のエリアセクション内カードを解析（店名・時給表示・data-rate）
  3. 突き合わせて差分レポートを C:/仕事/outputs/store_sync/ に保存
     - 新規（xlsxにあってサイトに無い）→ 貼り付け用カードHTML付き
     - 時給変更 → --apply で自動反映（表示ラベル＋data-rate）
     - 掲載落ち候補（サイトにあってxlsxに無い）→ レポートのみ（自動削除しない）
  4. 内部列（スカウトバック評価・内部備考）は読み飛ばす。レポートにも出さない
"""
import argparse
import datetime
import json
import re
import subprocess
import sys
from pathlib import Path

import openpyxl

SITE = Path(__file__).resolve().parent.parent
INDEX = SITE / "stores" / "index.html"
XLSX = Path(r"C:\仕事\context\一次情報\店舗情報まとめ最新.xlsx")
SLUG_MAP = Path(r"C:\仕事\outputs\xlsx_to_slug.json")
REPORT_DIR = Path(r"C:\仕事\outputs\store_sync")
# xlsx店名 → サイトカード店名 の表記ゆれ辞書。誤マッチ/取りこぼしが出たらここに足す
ALIASES = Path(r"C:\仕事\outputs\store_sync\name_aliases.json")

# xlsxのエリア名 → index.html のサブエリア見出しの接頭辞（「○○エリア 全店舗一覧」）
def area_short(area):
    a = re.sub(r"[（(].*?[)）]", "", area or "").strip()
    return a


def parse_xlsx():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["店舗一覧"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h}
    stores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        area = row[idx["エリア"]]
        name = row[idx["店舗名"]]
        if not area or not name:
            continue
        def num(col):
            v = row[idx[col]]
            try:
                return int(v)
            except (TypeError, ValueError):
                return None
        stores.append({
            "area": str(area).strip(),
            "name": str(name).strip(),
            "trial": num("体入時給(円〜)"),
            "regular": num("本入時給(円〜)"),
        })
    return stores


def rate_of(store):
    """カード表示に使う時給 = 体入時給（サイトの表示基準）。無ければ本入"""
    return store["trial"] or store["regular"]


def band_of(rate):
    if not rate:
        return "low"
    if rate >= 5000:
        return "high"
    if rate >= 3000:
        return "mid"
    return "low"


def tokens(s):
    parts = re.split(r"[（()）/、]", s)
    out = []
    for p in parts:
        p = re.sub(r"[\s・\-—｜|]", "", p).replace("元", "").replace("元:", "").upper().strip()
        if len(p) >= 2:
            out.append(p)
    return out


def match_score(xname, sname):
    """xlsx店名とサイトカード店名の一致度。>=44で候補扱い"""
    xt, st = tokens(xname), tokens(sname)
    # 1〜2文字の店名（蒼・響など）はトークンに乗らないので正規化全文一致で拾う
    def norm(s):
        return re.sub(r"[（(].*?[)）]|[\s・\-—｜|]", "", s).upper()
    if norm(xname) and norm(xname) == norm(sname):
        return 100
    best = 0
    for a in xt:
        for b in st:
            if a == b:
                best = max(best, 100)
            elif len(a) >= 4 and len(b) >= 4 and (a in b or b in a):
                best = max(best, 80)
            else:
                # 共通接頭辞4文字以上
                n = 0
                for ca, cb in zip(a, b):
                    if ca != cb:
                        break
                    n += 1
                if n >= 4:
                    best = max(best, 40 + n)
    return best


def parse_index():
    """area-* セクション内のカードを (name, rate, band, slug, subarea, start, end) で返す"""
    html = INDEX.read_text(encoding="utf-8")
    sections = [(m.start(), m.group(1)) for m in
                re.finditer(r'<div id="(area-[a-z0-9]+)" class="area-section"', html)]
    if not sections:
        sys.exit("area-section が見つからない。index.htmlの構造が変わった可能性")
    sections.append((len(html), "END"))

    cards = []
    for (s, sec_id), (e, _) in zip(sections, sections[1:]):
        seg = html[s:e]
        # サブエリア見出し位置（「○○エリア 全店舗一覧」）
        heads = [(m.start(), re.sub(r"エリア.*", "", m.group(1)).strip())
                 for m in re.finditer(r'<h3 class="text-sm[^"]*">([^<]+)</h3>', seg)]
        card_starts = [m.start() for m in re.finditer(r'<div class="store-card', seg)]
        card_starts.append(len(seg))
        for cs, ce in zip(card_starts, card_starts[1:]):
            chunk = seg[cs:ce]
            m_name = re.search(r"<h3[^>]*>([^<]+)</h3>", chunk)
            if not m_name:
                continue
            name = m_name.group(1).replace("⭐", "").strip()
            m_rate = re.search(r"時給([\d,]+)円〜", chunk)
            rate = int(m_rate.group(1).replace(",", "")) if m_rate else None
            m_band = re.search(r'data-rate="(\w+)"', chunk)
            m_slug = re.search(r'href="/stores/([\w\-]+)\.html"', chunk)
            subarea = ""
            for hpos, hname in heads:
                if hpos < cs:
                    subarea = hname
            cards.append({
                "name": name,
                "rate": rate,
                "band": m_band.group(1) if m_band else None,
                "slug": m_slug.group(1) if m_slug else None,
                "section": sec_id,
                "subarea": subarea,
                "abs_start": s + cs,
                "abs_end": s + ce,
            })
    return html, cards


def load_slug_map():
    if not SLUG_MAP.exists():
        return {}
    data = json.loads(SLUG_MAP.read_text(encoding="utf-8"))
    return {(r["xlsx_name"], r["area"]): r["slug"] for r in data.get("matched", [])}


def load_aliases():
    if not ALIASES.exists():
        return {}
    return json.loads(ALIASES.read_text(encoding="utf-8"))


def build_matches(xstores, cards, slug_map, aliases):
    """xlsx店舗 ⇔ サイトカード の1対1マッチング"""
    pairs = []  # (score, xi, ci)
    for xi, xs in enumerate(xstores):
        x_sub = area_short(xs["area"])
        slug = slug_map.get((xs["name"], xs["area"]))
        alias = aliases.get(xs["name"])
        for ci, c in enumerate(cards):
            if alias and c["name"] == alias:
                pairs.append((300, xi, ci))
                continue
            # サブエリアが分かるカードはエリア一致を要求
            # （旧スラグ対応表に別エリア流用があるため、スラグ一致でもエリア違いは不可）
            if c["subarea"] and c["subarea"] != x_sub:
                continue
            # スラグ一致は最優先
            if slug and c["slug"] == slug:
                pairs.append((200, xi, ci))
                continue
            sc = match_score(xs["name"], c["name"])
            if sc >= 44:
                pairs.append((sc, xi, ci))
    pairs.sort(reverse=True)
    x_used, c_used = {}, {}
    for sc, xi, ci in pairs:
        if xi in x_used or ci in c_used:
            continue
        x_used[xi] = (ci, sc)
        c_used[ci] = xi
    return x_used, c_used


def card_snippet(xs):
    rate = rate_of(xs)
    band = band_of(rate)
    label = f"時給{rate:,}円〜" if rate else "時給要相談"
    return f'''          <!-- {xs["name"]} -->
          <div class="store-card min-w-0 bg-[#1a1a1a] rounded-xl p-4 border border-white/10 flex flex-col gap-2" data-rate="{band}">
            <div class="flex items-start justify-between gap-1 mb-0.5">
              <h3 class="min-w-0 text-sm font-bold text-white leading-snug break-all">{xs["name"]}</h3>
            </div>
            <div class="flex flex-wrap gap-1">
              <span class="text-[11px] bg-white/5 text-gray-400 px-2 py-0.5 rounded-full whitespace-nowrap">{label}</span>
            </div>
            <div class="flex gap-1.5 mt-auto pt-1">
              <a href="/#cta" class="flex-1 text-center text-xs text-white bg-brand-line hover:bg-brand-lineDark py-2.5 rounded-lg font-bold transition-colors">相談</a>
            </div>
          </div>'''


def apply_rate_changes(html, changes):
    """changes: [(card, new_rate)] を後ろから順に置換（位置ずれ防止）"""
    for card, new_rate in sorted(changes, key=lambda t: -t[0]["abs_start"]):
        chunk = html[card["abs_start"]:card["abs_end"]]
        new_label = f"時給{new_rate:,}円〜"
        new_chunk = re.sub(r"時給[\d,]+円〜", new_label, chunk, count=1)
        new_chunk = re.sub(r'data-rate="\w+"', f'data-rate="{band_of(new_rate)}"', new_chunk, count=1)
        html = html[:card["abs_start"]] + new_chunk + html[card["abs_end"]:]
    return html


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="時給変更をindex.htmlに反映")
    ap.add_argument("--commit", action="store_true", help="反映後にgit commit/push")
    args = ap.parse_args()

    xstores = parse_xlsx()
    html, cards = parse_index()
    slug_map = load_slug_map()
    x_used, c_used = build_matches(xstores, cards, slug_map, load_aliases())

    new_stores, rate_changes, weak_matches = [], [], []
    for xi, xs in enumerate(xstores):
        if xi not in x_used:
            new_stores.append(xs)
            continue
        ci, sc = x_used[xi]
        card = cards[ci]
        if sc < 80:
            weak_matches.append((xs, card, sc))
        xrate = rate_of(xs)
        if xrate and card["rate"] and xrate != card["rate"]:
            rate_changes.append((xs, card, xrate))
    dropped = [c for ci, c in enumerate(cards) if ci not in c_used]

    today = datetime.date.today().isoformat()
    lines = [f"# 店舗同期レポート {today}", ""]
    lines.append(f"xlsx: {len(xstores)}店舗 / サイトカード: {len(cards)}枚 / マッチ: {len(x_used)}件")
    lines.append("")

    lines.append(f"## 時給変更 {len(rate_changes)}件" + ("（--applyで反映済み）" if args.apply else "（未反映。--applyで反映）"))
    for xs, card, xrate in sorted(rate_changes, key=lambda t: t[0]["area"]):
        old = f"{card['rate']:,}" if card["rate"] else "?"
        lines.append(f"- [{area_short(xs['area'])}] {card['name']}: 時給{old}円〜 → 時給{xrate:,}円〜")
    lines.append("")

    lines.append(f"## 新規（サイト未掲載） {len(new_stores)}件")
    lines.append("エリアのカード欄に下のスニペットを貼れば掲載できる（見出しの店舗数カウントは手で+1）。")
    by_area = {}
    for xs in new_stores:
        by_area.setdefault(xs["area"], []).append(xs)
    for area in sorted(by_area):
        lines.append(f"\n### {area}（{len(by_area[area])}件）")
        for xs in by_area[area]:
            r = rate_of(xs)
            lines.append(f"- {xs['name']}（時給{r:,}〜）" if r else f"- {xs['name']}（時給不明）")
        lines.append("\n```html")
        for xs in by_area[area]:
            lines.append(card_snippet(xs))
        lines.append("```")
    lines.append("")

    lines.append(f"## 掲載落ち候補（xlsxに見当たらない） {len(dropped)}件")
    lines.append("店名表記ゆれの可能性もあるので自動削除はしない。閉店・移転ならカードを手動削除。")
    for c in sorted(dropped, key=lambda c: (c["section"], c["subarea"])):
        lines.append(f"- [{c['subarea'] or c['section']}] {c['name']}")
    lines.append("")

    if weak_matches:
        lines.append(f"## 要確認（あいまい一致） {len(weak_matches)}件")
        lines.append("スコア低めの名寄せ。誤マッチなら outputs/xlsx_to_slug.json に正しい対応を足す。")
        for xs, card, sc in weak_matches:
            lines.append(f"- xlsx「{xs['name']}」 ⇔ サイト「{card['name']}」（score {sc}）")
        lines.append("")

    if args.apply and rate_changes:
        html = apply_rate_changes(html, [(card, xrate) for _, card, xrate in rate_changes])
        INDEX.write_text(html, encoding="utf-8")
        lines.append(f"> --apply: {len(rate_changes)}件の時給をindex.htmlに反映した。")

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report_path = REPORT_DIR / f"店舗同期レポート_{today}.md"
    report_path.write_text("\n".join(lines), encoding="utf-8")

    print(f"レポート: {report_path}")
    print(f"時給変更 {len(rate_changes)} / 新規 {len(new_stores)} / 掲載落ち候補 {len(dropped)} / 要確認 {len(weak_matches)}")

    if args.apply and args.commit and rate_changes:
        subprocess.run(["git", "add", "stores/index.html"], cwd=SITE, check=True)
        msg = f"店舗時給を最新xlsxに同期（{len(rate_changes)}件・{today}）"
        subprocess.run(["git", "commit", "-m", msg], cwd=SITE, check=True)
        subprocess.run(["git", "push"], cwd=SITE, check=True)
        print("git push 完了")


if __name__ == "__main__":
    main()
